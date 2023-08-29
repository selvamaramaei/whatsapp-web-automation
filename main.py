 
from datetime import  datetime
import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from openpyxl import Workbook , load_workbook
import time



driver_path = "C:\\Users\\90534\\Documents\\Python-projeleri\\selenium-whatsapp\\chromedriver-win64\\chromedriver.exe"
driver = webdriver.Chrome()
URL = "C:\\Users\\90534\\OneDrive\\Belgeler\\whatsapp_data.xlsx"




def find_website(url):
    driver.get(url)
    
 
def login(number):

    login = driver.find_element(By.CSS_SELECTOR, "._2rQUO span")
    login.click() 

    time.sleep(3)

    phone_number = driver.find_element(By.CSS_SELECTOR , ".selectable-text ")
    phone_number.send_keys(number)

    time.sleep(2)

    nextBtn = driver.find_element(By.XPATH , "//*/div/div/div[3]/div[1]/div/div[3]/div[3]/div/div/div")
    nextBtn.click()

    time.sleep(25)



def find_number(PhoneNumber):
    search_input = driver.find_element(By.CLASS_NAME , "selectable-text")
    search_input.send_keys(f"{PhoneNumber}")
    search_input.click()
    search_input.send_keys(Keys.ENTER)



def send_message(content , userinfo) :
    type = driver.find_element(By.CSS_SELECTOR, "._1VZX7 p")
    type.send_keys(f"Dear {userinfo} , {content}")
    type.send_keys(Keys.ENTER)



def load_df(URL):
    df = pd.read_excel(URL)
    return df



def query_data(df):
    wb = load_workbook(URL)
    ws = wb.active
    message_counter= 0
    for _ , row in df.iterrows():
        if pd.notnull(row["phone"]) and isinstance(row["phone"],str):
            try : 
                find_number(row["phone"])
                time.sleep(2)

                send_message("Hello",row["name-surname"])
                time.sleep(5)
                
                
                # Get the column indexes using column names
                has_send_col_idx = df.columns.get_loc("has_send")
                sending_date_col_idx = df.columns.get_loc("sending_date")

                # Update the cell values for the current row
                ws.cell(row=_ + 2, column=has_send_col_idx + 1, value="yes")
                sending_time = datetime.now()
                ws.cell(row=_+ 2, column=sending_date_col_idx + 1, value=sending_time)
            except:

                has_send_col_idx = df.columns.get_loc("has_send")
                ws.cell(row=_ + 2, column=has_send_col_idx + 1, value="No")
                time.sleep(2)
                sending_time = datetime.now()
                sending_date_col_idx = df.columns.get_loc("sending_date")
                ws.cell(row=_ + 2, column=sending_date_col_idx + 1, value=sending_time)       
    wb.save(URL)
       


def main():
    find_website("https://web.whatsapp.com/")
    time.sleep(10)
    login("*************")
    time.sleep(5)
    df = load_df(URL)
    result = query_data(df)
    print(result)
    input("press any key")



if __name__ == "__main__":
    main()


    

                   